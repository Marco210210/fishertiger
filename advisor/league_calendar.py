"""Convert legacy league fixture spreadsheets into the canonical calendar format.

This module is intentionally independent of the projection pipeline.  Its public
API is :func:`preprocess_legacy_calendar` for files and
:func:`parse_legacy_two_block_frame` for already-loaded pandas DataFrames.
"""
from __future__ import annotations

import argparse
from io import BytesIO
import json
import re
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook

SCHEMA_VERSION = "1.0"
_LEAGUE_DAY = re.compile(r"(\d+)ª\s+Giornata\s+lega", re.IGNORECASE)
_SERIE_A_DAY = re.compile(r"(\d+)ª\s+Giornata\s+serie\s+a", re.IGNORECASE)


def _text(value: object) -> str | None:
    if pd.isna(value):
        return None
    value = str(value).strip()
    return value or None


def _headers(frame: pd.DataFrame, start: int) -> list[tuple[int, int, int]]:
    """Return legacy headers from one fixture block as (row, league, Serie A)."""
    headers = []
    for row in range(len(frame)):
        league = _LEAGUE_DAY.search(_text(frame.iat[row, start]) or "")
        if not league:
            continue
        serie_a = _SERIE_A_DAY.search(_text(frame.iat[row, start + 2]) or "")
        if not serie_a:
            raise ValueError(
                f"legacy calendar: Serie A matchday missing beside league matchday {league.group(1)}"
            )
        headers.append((row, int(league.group(1)), int(serie_a.group(1))))
    return headers


def validate_calendar(calendar: Mapping[str, Any]) -> None:
    """Validate structural invariants without assuming a league size or format."""
    required = {"schema_version", "league_id", "teams", "participants_count", "matchdays"}
    missing = required - set(calendar)
    if missing:
        raise ValueError(f"calendar: missing required fields {sorted(missing)}")
    if calendar["schema_version"] != SCHEMA_VERSION:
        raise ValueError(f"calendar: unsupported schema version {calendar['schema_version']!r}")
    if not isinstance(calendar["league_id"], str) or not calendar["league_id"].strip():
        raise ValueError("calendar: league_id must be a non-empty string")
    teams = calendar["teams"]
    if not isinstance(teams, list) or not teams or any(not isinstance(team, str) or not team.strip() for team in teams):
        raise ValueError("calendar: teams must be a non-empty list of names")
    if len(set(teams)) != len(teams):
        raise ValueError("calendar: teams must be unique")
    if calendar["participants_count"] != len(teams):
        raise ValueError("calendar: participants_count must match teams")

    if not isinstance(calendar["matchdays"], list):
        raise ValueError("calendar: matchdays must be a list")
    numbers = set()
    known_teams = set(teams)
    for matchday in calendar["matchdays"]:
        if not isinstance(matchday, Mapping):
            raise ValueError("calendar: each matchday must be an object")
        number = matchday.get("number")
        serie_a_matchday = matchday.get("serie_a_matchday")
        fixtures = matchday.get("fixtures")
        if not isinstance(number, int) or number < 1 or number in numbers:
            raise ValueError("calendar: matchday numbers must be positive and unique")
        if not isinstance(serie_a_matchday, int) or serie_a_matchday < 1:
            raise ValueError(f"calendar: matchday {number} has an invalid Serie A matchday")
        if not isinstance(fixtures, list) or not fixtures:
            raise ValueError(f"calendar: matchday {number} must contain fixtures")
        numbers.add(number)
        playing = set()
        for fixture in fixtures:
            if not isinstance(fixture, Mapping):
                raise ValueError(f"calendar: matchday {number} contains an invalid fixture")
            home, away = fixture.get("home"), fixture.get("away")
            if not isinstance(home, str) or not isinstance(away, str) or not home or not away:
                raise ValueError(f"calendar: matchday {number} has an incomplete fixture")
            if home == away:
                raise ValueError(f"calendar: matchday {number} contains a self fixture for {home!r}")
            if home not in known_teams or away not in known_teams:
                raise ValueError(f"calendar: matchday {number} references an unknown team")
            if home in playing or away in playing:
                raise ValueError(f"calendar: team appears more than once in matchday {number}")
            playing.update((home, away))


def parse_legacy_two_block_frame(frame: pd.DataFrame, league_id: str) -> dict[str, Any]:
    """Parse the current two-block Leghe Fantacalcio layout from a DataFrame.

    The left block starts at column 0 and the right block at column 6.  Fixture
    rows continue until the next header, so the number of fixtures is inferred.
    """
    if frame.shape[1] < 10:
        raise ValueError("legacy calendar: expected the two fixture blocks used by calendario_lega.xlsx")

    matchdays: dict[int, dict[str, Any]] = {}
    for start in (0, 6):
        block_headers = _headers(frame, start)
        for index, (row, number, serie_a_matchday) in enumerate(block_headers):
            if number in matchdays:
                raise ValueError(f"legacy calendar: duplicate league matchday {number}")
            end = block_headers[index + 1][0] if index + 1 < len(block_headers) else len(frame)
            fixtures = []
            for fixture_row in range(row + 1, end):
                home, away = _text(frame.iat[fixture_row, start]), _text(frame.iat[fixture_row, start + 3])
                if home is None and away is None:
                    continue
                if home is None or away is None:
                    raise ValueError(f"legacy calendar: incomplete fixture at league matchday {number}")
                fixtures.append({"home": home, "away": away})
            if not fixtures:
                raise ValueError(f"legacy calendar: no fixtures at league matchday {number}")
            matchdays[number] = {
                "number": number,
                "serie_a_matchday": serie_a_matchday,
                "fixtures": fixtures,
            }
    if not matchdays:
        raise ValueError("legacy calendar: no league matchday headers found")

    ordered_matchdays = [matchdays[number] for number in sorted(matchdays)]
    teams = sorted({team for day in ordered_matchdays for fixture in day["fixtures"] for team in (fixture["home"], fixture["away"])})
    calendar = {
        "schema_version": SCHEMA_VERSION,
        "league_id": league_id,
        "teams": teams,
        "participants_count": len(teams),
        "matchdays": ordered_matchdays,
    }
    validate_calendar(calendar)
    return calendar


def load_legacy_calendar(source: str | Path, sheet_name: str = "Calendario") -> pd.DataFrame:
    """Load a legacy spreadsheet, keeping file I/O separate from parsing."""
    try:
        return pd.read_excel(source, sheet_name=sheet_name, header=None)
    except ValueError as error:
        if "Worksheet named" in str(error) and "not found" in str(error):
            raise ValueError("legacy calendar: worksheet 'Calendario' is required") from error
        raise


def preprocess_legacy_calendar(source: str | Path, league_id: str, sheet_name: str = "Calendario") -> dict[str, Any]:
    """Read a legacy workbook and return its validated canonical calendar."""
    return parse_legacy_two_block_frame(load_legacy_calendar(source, sheet_name), league_id)


def _round_robin_fixtures(teams: list[str]) -> list[list[tuple[str, str]]]:
    """Build one complete round robin for an even number of teams."""
    rotation = teams[:]
    rounds = []
    for _ in range(len(teams) - 1):
        rounds.append([(rotation[index], rotation[-index - 1]) for index in range(len(teams) // 2)])
        rotation = [rotation[0], rotation[-1], *rotation[1:-1]]
    return rounds


def build_legacy_calendar_template(participants: int = 8, matchdays: int = 36) -> bytes:
    """Return a sanitized legacy-layout workbook suitable for the default profile."""
    if participants < 2 or participants % 2:
        raise ValueError("calendar template: participants must be an even number of at least 2")
    if matchdays < 1:
        raise ValueError("calendar template: matchdays must be positive")

    teams = [f"Squadra {number}" for number in range(1, participants + 1)]
    rounds = _round_robin_fixtures(teams)
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Istruzioni"
    instructions.append(["Modello calendario della lega"])
    instructions.append(["Compila il foglio Calendario senza rinominarlo."])
    instructions.append(["Le squadre nelle colonne A/D e G/J devono corrispondere al profilo."])
    instructions.append(["Ogni giornata usa i titoli 'Nª Giornata lega' e 'Mª Giornata serie a'."])
    calendar = workbook.create_sheet("Calendario")
    for matchday in range(1, matchdays + 1):
        start = 1 if matchday % 2 else 7
        row = ((matchday - 1) // 2) * (participants // 2 + 1) + 1
        calendar.cell(row, start, f"{matchday}ª Giornata lega")
        calendar.cell(row, start + 2, f"{matchday}ª Giornata serie a")
        for fixture_row, (home, away) in enumerate(rounds[(matchday - 1) % len(rounds)], start=row + 1):
            calendar.cell(fixture_row, start, home)
            calendar.cell(fixture_row, start + 3, away)
    calendar.freeze_panes = "A2"
    for column in ("A", "D", "G", "J"):
        calendar.column_dimensions[column].width = 20
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert calendario_lega.xlsx to canonical league-calendar JSON.")
    parser.add_argument("source", type=Path, help="Legacy calendario_lega.xlsx path")
    parser.add_argument("destination", type=Path, help="Output JSON path")
    parser.add_argument("--league-id", required=True, help="Stable identifier for this league")
    parser.add_argument("--sheet-name", default="Calendario", help="Workbook sheet name (default: Calendario)")
    args = parser.parse_args(argv)
    calendar = preprocess_legacy_calendar(args.source, args.league_id, args.sheet_name)
    args.destination.parent.mkdir(parents=True, exist_ok=True)
    args.destination.write_text(json.dumps(calendar, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
